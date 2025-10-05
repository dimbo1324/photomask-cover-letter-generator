from pathlib import Path
from datetime import datetime
from typing import Optional, Dict, Any
from openpyxl import load_workbook


class ExcelWorker:
    def __init__(
        self,
        src: Path,
        prefix: str,
        above_value: str,
        options: Optional[Dict[str, Any]] = None,
    ):
        self.src = Path(src)
        self.prefix = prefix
        self.above_value = above_value
        opts = options or {}
        self.search_text = opts.get("search_text", "№ слоя")
        self.search_col = int(opts.get("search_col", 5))
        self.source_col = int(opts.get("source_col", 6))
        self.target_col = int(opts.get("target_col", 7))
        self.dest_option = opts.get("dest", None)
        self.case_sensitive = bool(opts.get("case_sensitive", False))
        self.modified_rows = []
        if not self.src.exists():
            raise FileNotFoundError(f"Файл не найден: {self.src}")

    def _matches_search(self, cell_value: Any) -> bool:
        if cell_value is None:
            return False
        text = str(cell_value)
        if not self.case_sensitive:
            return self.search_text.lower() in text.lower()
        return self.search_text in text

    def _resolve_dest_path(self) -> Optional[Path]:
        if self.dest_option is None:
            return None
        dest = Path(self.dest_option)
        if str(self.dest_option).endswith(("/", "\\")) or dest.is_dir():
            return dest
        if dest.parent and str(dest.parent) != ".":
            if dest.suffix:
                return dest
            if dest.exists() and dest.is_dir():
                return dest
        if dest.name == str(self.dest_option):
            if dest.suffix:
                return self.src.parent / dest.name
            return self.src.parent / (dest.name + ".xlsx")
        return dest

    def run(
        self, save_copy: bool = True, inplace: bool = False
    ) -> (Optional[Path], list):
        wb = load_workbook(self.src)
        ws = wb.active
        self.modified_rows = []
        for row in range(1, ws.max_row + 1):
            cell_search = ws.cell(row=row, column=self.search_col).value
            if not self._matches_search(cell_search):
                continue
            cell_source_val = ws.cell(row=row, column=self.source_col).value
            if cell_source_val is None:
                continue
            s = str(cell_source_val).strip()
            if s == "":
                continue
            ws.cell(row=row, column=self.target_col).value = f"{self.prefix}{s}"
            if row > 1:
                ws.cell(row - 1, column=self.target_col).value = self.above_value
            self.modified_rows.append(row)
        saved_path = None
        if inplace:
            wb.save(self.src)
            saved_path = self.src
        elif save_copy:
            resolved = self._resolve_dest_path()
            if resolved is None:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                dest_path = self.src.with_name(f"{self.src.stem}_{timestamp}.xlsx")
            else:
                if resolved.exists() and resolved.is_dir():
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    dest_path = resolved / f"{self.src.stem}_{timestamp}.xlsx"
                else:
                    dest_path = resolved
            wb.save(dest_path)
            saved_path = dest_path
        else:
            saved_path = None
        return saved_path, self.modified_rows
